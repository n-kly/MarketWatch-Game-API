from bs4 import BeautifulSoup as soup
import requests
import openpyxl
from openpyxl.formatting.rule import *
from openpyxl.styles import *
from ast import literal_eval
from MWSetup import URL1

# Initialize all variables
dataArray = []
exPath = "Weekly\\Weekly Spreadsheet.xlsx"
loPath = "Weekly\\LOG\\"

try: URL = URL1
except NameError: print("Go into MWSetup and set a URL"); exit()

x = 0
status = 0
week = int(input('What week is it?: '))
pWeek = str(week - 1)
week = str(week)
ghettoKeys = []
removeList = []
counter = 0

def removeNonAscii(s): return "".join(i for i in s if ord(i)<126 and ord(i)>31) # Imported function to remove formatting

# Get the code of the webpage
def getCode(page): #Input the page of results to get code from
        pageIndex = (page*10)
        gameRankURL = (f'{URL}?partial=true&index={pageIndex}')

        s = requests.session() # Start session
        r = s.get(gameRankURL) # Get code from URL
        code = soup(r.content, 'html.parser') # Parse code for HTML
        return code # Returns HTML of rankings page

# Get the number of people in the game
code = getCode(0)
String = code.find("span",{"class":"text align--center"}).text
numberPeople = int(String[18:-8])

# Parse the code for the ranking data
def getTable(code): # Input code of the site
    table = code.find("table",{"class":"table table--primary ranking"}).tbody.findAll("tr",{"class":"table__row"}) # Gets code we want; Table is a list of rows
    return table # Returns the HTML table of the rankings page

# Create an array out of the ranking data
def rank(table,index): # Input HTML table; Input rank index
    row = table[index].findAll("td") # Row is rows of data
    rank = int(row[0].text) # Casts into int
    name = removeNonAscii(row[1].text) # Removes formatting
    trades = int(row[4].text) # Cast into int
    treturn = float(((row[5].text).replace("$","")).replace(",","")) # Removes the $ sign; Removes the ,; cast into float

    Person = [name,rank,0,trades,0,treturn,0]

    global status
    status += 1
    print(f"    Person added ({status}/{numberPeople})")
    return Person # Returns the statistics of the index

def getRank(page,index): return rank(getTable(getCode(page)),(index-(page*10))) # Simplify a compound of functions into a single one

print("Creating dataArray")# Attach together all of the ranking data into a 2d array

for i in range(numberPeople): # Loop for however many people there are
    page = i // 10
    dataArray.append(getRank(page,i))
print("")

dataArray.sort() # Sorts array alphabetically

f = open(f"{loPath}Week {str(pWeek)}.txt","r") # Opens previous weeks data into a variable

PdataArray = literal_eval(f.read())
f.close()
print("")

PdataArray.append(["Cheeky Patch"]) # Fixes a bug with index values

print("Cleansing dataArray")
# TODO use a dictionary instead of a list please
for i in range(len(dataArray)): # Find in new thats not in old; Add in new
    if dataArray[i][0] != PdataArray[i][0]: # Checks to see if anyone new has joined and gives them a PdataArray value
        List = [(dataArray[i][0]),0,0,0,0,0,0]
        PdataArray.insert(i,List)
        print(f"    {dataArray[i][0]}'s person entry added")
    else:
        pass
    dataArray[i][2] = dataArray[i][1] - PdataArray[i][1] # Sets movement values
    dataArray[i][4] = dataArray[i][3] - PdataArray[i][3] # Sets new trade values
    dataArray[i][6] = dataArray[i][5] - PdataArray[i][5] # Sets relative return values
print("")


for data in dataArray: # Creates a list of all "keys"
    ghettoKeys.append(data[0])

for i in range(len(PdataArray)): # Find in old thats not in new; Delete in old
    if PdataArray[i][0] not in ghettoKeys: # Checks to see if anyone has left the game
        removeList.append((i-counter)) # Creates list of indexes to remove
        counter += 1
    else:
        pass

for i in removeList:
    print(f"    {PdataArray[i][0]}'s person entry removed")
    PdataArray.pop(i)


print("Saving dataArray")# Logs the weeks data in a text file
f = open(f"{loPath}Week {week}.txt", "x")
f.write(str(dataArray))
f.close()
print("")

print("Creating excel sheet") # Create and load excel spreadsheet
wb = openpyxl.load_workbook(exPath) # Loads the excel spread sheet
wb.create_sheet(str(f"Week {week}")) # Creates worksheet
ws = wb[f"Week {week}"] # Picks active worksheet

# Initialize excel locations
row = 2
col = 1
Col = 1

print("    Writing headers")# Write headers
h = ws.cell(1,Col+0)
h.value = "Name"
h = ws.cell(1,Col+1)
h.value = "Rank"
h = ws.cell(1,Col+2)
h.value = "Movement"
h = ws.cell(1,Col+3)
h.value = "Trades"
h = ws.cell(1,Col+4)
h.value = "New trades"
h = ws.cell(1,Col+5)
h.value = "Total return"
h = ws.cell(1,Col+6)
h.value = "Weekly return"

print("    Writing data")# Write data into values
for data in dataArray:
    for dataRow in data:
        c = ws.cell(row,col)
        c.value = dataRow
        col += 1
    col = 1
    row += 1

cells = str(int(numberPeople)+1)

print("    Formatting")
# Adds formatting to all cols
# Adds formatting to rank
ws.conditional_formatting.add(f"B2:B{cells}",CellIsRule(operator="equal",formula=['1'], stopIfTrue=False, fill=(PatternFill(start_color="C9B037",end_color="C9B037",fill_type="solid")))) #Makes #1 gold
ws.conditional_formatting.add(f"B2:B{cells}",CellIsRule(operator="equal",formula=['2'], stopIfTrue=False, fill=(PatternFill(start_color="B4B4B4",end_color="B4B4B4",fill_type="solid")))) #Makes #2 silver
ws.conditional_formatting.add(f"B2:B{cells}",CellIsRule(operator="equal",formula=['3'], stopIfTrue=False, fill=(PatternFill(start_color="AD8A56",end_color="AD8A56",fill_type="solid")))) #Makes #3 bronze
ws.conditional_formatting.add(f"B2:B{cells}",CellIsRule(operator="equal",formula=[f'{numberPeople}'], stopIfTrue=False, fill=(PatternFill(start_color="004b90",end_color="004b90",fill_type="solid")))) #Makes last blue

# Adds formatting to movement
rule = IconSetRule('5Arrows', 'num', [-2,-2,0,1,3], showValue=None, percent=None, reverse=None)
ws.conditional_formatting.add(f"C2:C{cells}",rule)

#Adds formatting to total return
ws.conditional_formatting.add(f'F1:F{cells}',ColorScaleRule(start_type='percent', start_value=10, start_color='FF0000',mid_type='num', mid_value=0, mid_color='FFFFFF',end_type='percent', end_value=90, end_color='00FF00'))

#Adds formatting to weekly return
ws.conditional_formatting.add(f"G2:G{cells}",CellIsRule(operator="equal",formula=[f'LARGE(G$2:G${cells},1)'], stopIfTrue=False, fill=(PatternFill(start_color="C9B037",end_color="C9B037",fill_type="solid")))) #Makes #1 gold
ws.conditional_formatting.add(f"G2:G{cells}",CellIsRule(operator="equal",formula=[f'LARGE(G$2:G${cells},2)'], stopIfTrue=False, fill=(PatternFill(start_color="B4B4B4",end_color="B4B4B4",fill_type="solid")))) #Makes #2 silver
ws.conditional_formatting.add(f"G2:G{cells}",CellIsRule(operator="equal",formula=[f'LARGE(G$2:G${cells},3)'], stopIfTrue=False, fill=(PatternFill(start_color="AD8A56",end_color="AD8A56",fill_type="solid")))) #Makes #3 bronze
ws.conditional_formatting.add(f"G2:G{cells}",CellIsRule(operator="equal",formula=[f'SMALL(G$2:G${cells},1)'], stopIfTrue=False, fill=(PatternFill(start_color="004b90",end_color="004b90",fill_type="solid")))) #Makes last blue

# Save and close excel file
wb.save(exPath)
wb.close()
print("")
print("Done")

input("Press any key to exit") 
